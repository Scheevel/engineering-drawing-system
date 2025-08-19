from typing import List, Optional, Dict, Any
from sqlalchemy.orm import Session, joinedload
import uuid
import tempfile
import os
import logging
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.database import Component, Drawing, Project, Dimension, Specification
from app.models.export import ExportRequest, ExportFormat
from app.core.config import settings

logger = logging.getLogger(__name__)

class ExportService:
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    async def export_data(self, request: ExportRequest, db: Session) -> str:
        """Export component data in the specified format"""
        try:
            # Get component data
            components_data = await self._get_components_data(request.component_ids, request, db)
            
            if request.format == ExportFormat.EXCEL:
                return await self._export_to_excel(components_data, request)
            elif request.format == ExportFormat.CSV:
                return await self._export_to_csv(components_data, request)
            else:
                raise ValueError(f"Unsupported export format: {request.format}")
                
        except Exception as e:
            logger.error(f"Error exporting data: {str(e)}")
            raise
    
    async def _get_components_data(
        self, 
        component_ids: List[str], 
        request: ExportRequest, 
        db: Session
    ) -> List[Dict[str, Any]]:
        """Get component data with related information"""
        try:
            # Build query with eager loading
            query = db.query(Component).options(
                joinedload(Component.drawing).joinedload(Drawing.project),
                joinedload(Component.dimensions) if request.include_dimensions else None,
                joinedload(Component.specifications) if request.include_specifications else None
            ).filter(Component.id.in_(component_ids))
            
            components = query.all()
            
            # Convert to export format
            export_data = []
            for component in components:
                data = {
                    "piece_mark": component.piece_mark,
                    "component_type": component.component_type,
                    "description": component.description,
                    "quantity": component.quantity,
                    "material_type": component.material_type,
                    "drawing_file": component.drawing.file_name,
                    "sheet_number": component.drawing.sheet_number,
                    "project_name": component.drawing.project.name if component.drawing.project else "Unassigned",
                    "confidence_score": component.confidence_score,
                    "created_at": component.created_at.isoformat() if component.created_at else None
                }
                
                # Add dimensions if requested
                if request.include_dimensions and component.dimensions:
                    for i, dim in enumerate(component.dimensions):
                        data[f"dimension_{i+1}_type"] = dim.dimension_type
                        data[f"dimension_{i+1}_value"] = dim.nominal_value
                        data[f"dimension_{i+1}_unit"] = dim.unit
                        data[f"dimension_{i+1}_tolerance"] = dim.tolerance
                
                # Add specifications if requested
                if request.include_specifications and component.specifications:
                    for i, spec in enumerate(component.specifications):
                        data[f"spec_{i+1}_type"] = spec.specification_type
                        data[f"spec_{i+1}_value"] = spec.value
                        data[f"spec_{i+1}_description"] = spec.description
                
                export_data.append(data)
            
            return export_data
            
        except Exception as e:
            logger.error(f"Error getting components data: {str(e)}")
            raise
    
    async def _export_to_excel(self, data: List[Dict[str, Any]], request: ExportRequest) -> str:
        """Export data to Excel format"""
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Components"
            
            if not data:
                # Create empty file
                file_path = os.path.join(self.temp_dir, f"export_{uuid.uuid4()}.xlsx")
                wb.save(file_path)
                return file_path
            
            # Get all column names
            all_columns = set()
            for row in data:
                all_columns.update(row.keys())
            columns = sorted(list(all_columns))
            
            # Write headers
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for col_idx, column in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=column.replace("_", " ").title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, column in enumerate(columns, 1):
                    value = row_data.get(column, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            file_path = os.path.join(self.temp_dir, f"export_{uuid.uuid4()}.xlsx")
            wb.save(file_path)
            
            logger.info(f"Excel export created: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error creating Excel export: {str(e)}")
            raise
    
    async def _export_to_csv(self, data: List[Dict[str, Any]], request: ExportRequest) -> str:
        """Export data to CSV format"""
        try:
            if not data:
                # Create empty CSV
                file_path = os.path.join(self.temp_dir, f"export_{uuid.uuid4()}.csv")
                pd.DataFrame().to_csv(file_path, index=False)
                return file_path
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Save to CSV
            file_path = os.path.join(self.temp_dir, f"export_{uuid.uuid4()}.csv")
            df.to_csv(file_path, index=False)
            
            logger.info(f"CSV export created: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error creating CSV export: {str(e)}")
            raise
    
    async def generate_pdf_report(
        self,
        component_ids: List[str],
        project_id: Optional[str] = None,
        include_images: bool = False,
        db: Session = None
    ) -> str:
        """Generate a PDF report for selected components"""
        try:
            # For now, create a simple text file as placeholder
            # In a real implementation, this would use a PDF library like ReportLab
            
            components = db.query(Component).options(
                joinedload(Component.drawing).joinedload(Drawing.project),
                joinedload(Component.dimensions),
                joinedload(Component.specifications)
            ).filter(Component.id.in_(component_ids)).all()
            
            # Create report content
            report_lines = [
                "ENGINEERING DRAWING COMPONENT REPORT",
                "=" * 50,
                f"Generated: {datetime.now().isoformat()}",
                f"Components: {len(components)}",
                "",
            ]
            
            for component in components:
                report_lines.extend([
                    f"Piece Mark: {component.piece_mark}",
                    f"Type: {component.component_type}",
                    f"Description: {component.description or 'N/A'}",
                    f"Quantity: {component.quantity}",
                    f"Drawing: {component.drawing.file_name}",
                    f"Project: {component.drawing.project.name if component.drawing.project else 'N/A'}",
                    ""
                ])
                
                if component.dimensions:
                    report_lines.append("Dimensions:")
                    for dim in component.dimensions:
                        report_lines.append(f"  - {dim.dimension_type}: {dim.nominal_value} {dim.unit}")
                    report_lines.append("")
                
                if component.specifications:
                    report_lines.append("Specifications:")
                    for spec in component.specifications:
                        report_lines.append(f"  - {spec.specification_type}: {spec.value}")
                    report_lines.append("")
                
                report_lines.append("-" * 30)
                report_lines.append("")
            
            # Save to file
            file_path = os.path.join(self.temp_dir, f"report_{uuid.uuid4()}.txt")
            with open(file_path, 'w') as f:
                f.write("\\n".join(report_lines))
            
            logger.info(f"PDF report created: {file_path}")
            return file_path
            
        except Exception as e:
            logger.error(f"Error creating PDF report: {str(e)}")
            raise
    
    async def list_templates(self) -> List[Dict[str, Any]]:
        """List available export templates"""
        return [
            {
                "name": "standard",
                "description": "Standard component export with basic fields",
                "fields": ["piece_mark", "component_type", "quantity", "drawing_file"]
            },
            {
                "name": "detailed", 
                "description": "Detailed export with dimensions and specifications",
                "fields": ["piece_mark", "component_type", "quantity", "dimensions", "specifications"]
            },
            {
                "name": "summary",
                "description": "Summary export with aggregated data",
                "fields": ["piece_mark", "component_type", "quantity"]
            }
        ]